from flask import Blueprint, request, jsonify, send_file
from utils.auth_utils import require_auth
from db import DatabaseConnection
from psycopg2.extras import RealDictCursor
from utils.crypto_utils import decrypt_data
import openpyxl
from openpyxl.styles import Font, PatternFill
import io
from datetime import datetime

admin_bp = Blueprint('admin', __name__)

@admin_bp.route('/organizers/pending', methods=['GET'])
@require_auth(roles=['admin'])
def get_pending_organizers(current_user):
    """
    Get a list of all organizers that are currently in 'pending' status for THIS admin's club.
    """
    try:
        club_id = current_user.get('club_id')
        if not club_id:
             return jsonify([]), 200 # Should not happen for admin but safety first

        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT u.id, u.full_name, u.email, u.reg_no, u.phone_number, c.name as organization_name, u.dob, u.address, u.created_at
                    FROM users u
                    LEFT JOIN clubs c ON u.club_id = c.id
                    WHERE u.role = 'organizer' 
                    AND u.account_status = 'pending'
                    AND u.club_id = %s 
                    ORDER BY u.created_at DESC
                """, (club_id,))
                organizers = cur.fetchall()
                
                # Decrypt only personal fields (orgName is plaintext now)
                result = []
                for row in organizers:
                    org = dict(row)
                    org['phone_number'] = decrypt_data(org['phone_number'])
                    org['dob'] = decrypt_data(org['dob'])
                    org['address'] = decrypt_data(org['address'])
                    result.append(org)
                    
                return jsonify(result), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
@admin_bp.route('/organizers/active', methods=['GET'])
@require_auth(roles=['admin'])
def get_active_organizers(current_user):
    """
    Get a list of all organizers that are currently approved ('active') for THIS admin's club.
    """
    try:
        club_id = current_user.get('club_id')
        if not club_id:
             return jsonify([]), 200

        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT u.id, u.full_name, u.email, u.reg_no, u.phone_number, c.name as organization_name, u.created_at
                    FROM users u
                    LEFT JOIN clubs c ON u.club_id = c.id
                    WHERE u.role = 'organizer' 
                    AND u.account_status = 'active'
                    AND u.club_id = %s 
                    ORDER BY u.full_name ASC
                """, (club_id,))
                organizers = cur.fetchall()
                
                result = []
                for row in organizers:
                    org = dict(row)
                    org['phone_number'] = decrypt_data(org['phone_number'])
                    result.append(org)
                    
                return jsonify(result), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@admin_bp.route('/organizers/<int:user_id>/<action>', methods=['POST'])
@require_auth(roles=['admin'])
def process_organizer(current_user, user_id, action):
    """
    Approve or reject an organizer account.
    Works for both new applicants and existing members (revocation).
    """
    if action not in ['approve', 'reject']:
        return jsonify({"error": "Invalid action parameter"}), 400
        
    new_status = 'active' if action == 'approve' else 'rejected'
    club_id = current_user.get('club_id')
    if not club_id:
        return jsonify({"error": "Admin has no assigned club."}), 403
    
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    UPDATE users 
                    SET account_status = %s 
                    WHERE id = %s 
                    AND role = 'organizer' 
                    AND club_id = %s 
                    RETURNING id, email, full_name, club_id
                """, (new_status, user_id, club_id))
                
                updated = cur.fetchone()
                
                if not updated:
                    return jsonify({"error": "Organizer not found or belongs to another organization"}), 404
                    
                # Get the actual club name safely
                cur.execute("SELECT name FROM clubs WHERE id = %s", (updated['club_id'],))
                club_row = cur.fetchone()
                club_name_actual = club_row['name'] if club_row else 'Unknown Club'
                
                # Send Notification Email
                from utils.email_utils import send_organizer_status_email
                send_organizer_status_email(updated['email'], updated['full_name'], new_status, club_name_actual)
                    
                conn.commit()
                return jsonify({"message": f"Organizer successfully {new_status}"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- EVENT MANAGEMENT FOR ADMINS ---

@admin_bp.route('/events/pending', methods=['GET'])
@require_auth(roles=['admin'])
def get_pending_events(current_user):
    """Fetch all pending event requests for the admin's organization."""
    try:
        club_id = current_user.get('club_id')
        if not club_id:
            return jsonify([]), 200

        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT e.*, h.name as hall_name, u.full_name as organizer_name, c.name as club_name
                    FROM events e 
                    LEFT JOIN halls h ON e.hall_id = h.id
                    LEFT JOIN users u ON e.organizer_id = u.id
                    LEFT JOIN clubs c ON e.club_id = c.id
                    WHERE e.status = 'pending'
                    AND e.club_id = %s 
                    ORDER BY e.created_at ASC
                """, (club_id,))
                
                events = cur.fetchall()
                
                # Check Hall Availability for each event
                for event in events:
                    # Check if ANY approved event is already in this hall for these exact dates (overlap check)
                    cur.execute("""
                        SELECT id FROM events 
                        WHERE hall_id = %s 
                        AND status = 'approved'
                        AND id != %s
                        AND (
                            (start_date <= %s AND end_date >= %s) OR
                            (start_date <= %s AND end_date >= %s) OR
                            (start_date >= %s AND end_date <= %s)
                        )
                    """, (
                        event['hall_id'], event['id'], 
                        event['start_date'], event['start_date'],
                        event['end_date'], event['end_date'],
                        event['start_date'], event['end_date']
                    ))
                    conflict = cur.fetchone()
                    event['hall_available'] = (conflict is None)
                    
                    # If occupied, suggest other halls free for THESE dates
                    event['alternative_halls'] = []
                    if conflict:
                        cur.execute("""
                            SELECT name FROM halls 
                            WHERE id NOT IN (
                                SELECT hall_id FROM events 
                                WHERE status = 'approved'
                                AND (
                                    (start_date <= %s AND end_date >= %s) OR
                                    (start_date <= %s AND end_date >= %s) OR
                                    (start_date >= %s AND end_date <= %s)
                                )
                            )
                        """, (
                            event['start_date'], event['start_date'],
                            event['end_date'], event['end_date'],
                            event['start_date'], event['end_date']
                        ))
                        alts = cur.fetchall()
                        event['alternative_halls'] = [a['name'] for a in alts]
                    
                return jsonify(events), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@admin_bp.route('/events/approved', methods=['GET'])
@require_auth(roles=['admin'])
def get_approved_events_for_admin(current_user):
    """Fetch all approved event requests for the admin's organization."""
    try:
        club_id = current_user.get('club_id')
        if not club_id:
            return jsonify([]), 200

        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT e.*, h.name as hall_name, u.full_name as organizer_name, 
                           c.name as club_name, a.full_name as approved_by_name
                    FROM events e 
                    LEFT JOIN halls h ON e.hall_id = h.id
                    LEFT JOIN users u ON e.organizer_id = u.id
                    LEFT JOIN users a ON e.approved_by = a.id
                    LEFT JOIN clubs c ON e.club_id = c.id
                    WHERE e.status = 'approved'
                    AND e.club_id = %s 
                    ORDER BY e.start_date ASC
                """, (club_id,))
                
                events = cur.fetchall()
                # ISO format dates for JS
                for e in events:
                    if e['start_date']: e['start_date'] = e['start_date'].isoformat()
                    if e['end_date']: e['end_date'] = e['end_date'].isoformat()
                    if e['reg_deadline']: e['reg_deadline'] = e['reg_deadline'].isoformat()
                    if e['created_at']: e['created_at'] = e['created_at'].isoformat()
                    
                return jsonify(events), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@admin_bp.route('/clubs', methods=['GET'])
@require_auth(roles=['admin'])
def get_admin_clubs(current_user):
    """Fetch the list of clubs this admin is responsible for."""
    try:
        club_name = current_user.get('orgName', '')
        clubs = [c.strip() for c in club_name.split(',') if c.strip()]
        return jsonify(clubs), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@admin_bp.route('/events/<int:event_id>/process', methods=['POST'])
@require_auth(roles=['admin'])
def process_event(current_user, event_id):
    """Approve or reject an event with a message."""
    data = request.json
    action = data.get('action') # 'approve' or 'reject'
    message = data.get('message', "")
    
    if action not in ['approved', 'rejected']:
        return jsonify({"error": "Invalid action"}), 400
        
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Security: Ensure this event belongs to the specific club ID this admin manages
                club_id = current_user.get('club_id')
                if not club_id:
                    return jsonify({"error": "Admin has no assigned club."}), 403
                
                cur.execute("""
                    SELECT e.id, e.reg_amount, c.razorpay_key_id, c.razorpay_key_secret 
                    FROM events e
                    LEFT JOIN clubs c ON e.club_id = c.id
                    WHERE e.id = %s AND e.club_id = %s
                """, (event_id, club_id))
                
                event_data = cur.fetchone()
                if not event_data:
                    return jsonify({"error": "Unauthorized to process events for this organization"}), 403

                # Gateway Check for Paid Events
                if action == 'approved' and event_data['reg_amount'] and event_data['reg_amount'] > 0:
                    if not event_data['razorpay_key_id'] or not event_data['razorpay_key_secret']:
                        return jsonify({"error": "Club Payment Integration Missing: Please configure your Razorpay Keys in the System Profile before authorizing paid events."}), 400

                cur.execute("""
                    UPDATE events 
                    SET status = %s, admin_message = %s, approved_by = %s 
                    WHERE id = %s
                """, (action, message, current_user['sub'], event_id))
                conn.commit()
                return jsonify({"message": f"Event has been {action}"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@admin_bp.route('/events/calendar', methods=['GET'])
@require_auth(roles=['admin'])
def get_event_calendar(current_user):
    """Fetch all events across all halls for the centralized calendar."""
    try:
        club_id = current_user.get('club_id')
        if not club_id: return jsonify([]), 200

        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # 1. Fetch Events
                cur.execute("""
                    SELECT e.id, e.title, e.start_date, e.end_date, e.status, h.name as hall_name
                    FROM events e
                    JOIN halls h ON e.hall_id = h.id
                    WHERE e.club_id = %s
                    ORDER BY e.start_date ASC
                """, (club_id,))
                events = cur.fetchall()

                # 2. ISO format and check for overlaps globally
                for e in events:
                    e['start_date'] = e['start_date'].isoformat()
                    e['end_date'] = e['end_date'].isoformat()
                    
                    # Internal Overlap Check
                    cur.execute("""
                        SELECT id FROM events 
                        WHERE hall_id = (SELECT hall_id FROM events WHERE id = %s)
                        AND status = 'approved'
                        AND id != %s
                        AND (
                            (start_date <= (SELECT start_date FROM events WHERE id = %s) AND end_date >= (SELECT start_date FROM events WHERE id = %s)) OR
                            (start_date <= (SELECT end_date FROM events WHERE id = %s) AND end_date >= (SELECT end_date FROM events WHERE id = %s))
                        )
                    """, (e['id'], e['id'], e['id'], e['id'], e['id'], e['id']))
                    e['has_conflict'] = (cur.fetchone() is not None)

                return jsonify(events), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- DATA EXPORT & REPORTS ---

@admin_bp.route('/reports/list-all', methods=['GET'])
@require_auth(roles=['admin'])
def list_reports_events(current_user):
    """List all events managed by this admin for reporting purposes."""
    try:
        club_id = current_user.get('club_id')
        if not club_id: return jsonify([]), 200
        
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT e.id, e.title, e.start_date, h.name as venue_name,
                           (SELECT COUNT(*) FROM registrations WHERE event_id = e.id) as total_registrations
                    FROM events e
                    JOIN halls h ON e.hall_id = h.id
                    WHERE e.club_id = %s
                    ORDER BY e.start_date DESC
                """, (club_id,))
                return jsonify(cur.fetchall()), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@admin_bp.route('/reports/export/<int:event_id>', methods=['GET'])
@require_auth(roles=['admin'])
def export_event_report(current_user, event_id):
    """Export a comprehensive XLSX report for a specific event."""
    try:
        club_id = current_user.get('club_id')
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # 1. Fetch Event & Venue Details
                cur.execute("""
                    SELECT e.*, h.name as venue_name, h.description as venue_description, h.capacity,
                           u.full_name as organizer_name, cl.name as club_name
                    FROM events e
                    LEFT JOIN halls h ON e.hall_id = h.id
                    LEFT JOIN users u ON e.organizer_id = u.id
                    LEFT JOIN clubs cl ON e.club_id = cl.id
                    WHERE e.id = %s AND e.club_id = %s
                """, (event_id, club_id))
                event = cur.fetchone()
                if not event: return jsonify({"error": "Event not found or unauthorized"}), 404

                cur.execute("""
                    SELECT u.full_name, u.reg_no, u.email, u.college_email, 
                           r.status as reg_status, r.registered_at, r.team_name,
                           COALESCE(a.manual_present, FALSE) as manual_attendance,
                           COALESCE(a.otp_present, FALSE) as otp_attendance
                    FROM registrations r
                    JOIN registration_members rm ON r.id = rm.registration_id
                    JOIN users u ON rm.student_id = u.id
                    LEFT JOIN attendance a ON r.event_id = a.event_id AND u.id = a.student_id
                    WHERE r.event_id = %s AND r.status = 'approved'
                    ORDER BY r.team_name ASC, u.reg_no ASC
                """, (event_id,))
                participants = cur.fetchall()

                # Generate Workbook
                wb = openpyxl.Workbook()
                
                # Sheet 1: Summary Dossier
                ws1 = wb.active
                ws1.title = "Event Summary"
                
                # Format JSON fields safely
                import json
                flow_str = json.dumps(event['event_flow'], indent=2) if event['event_flow'] else "Standard Flow"
                refreshments_str = json.dumps(event['refreshments'], indent=2) if event['refreshments'] else "No catering"

                summary_data = [
                    ["EVENT REPORT DOSSIER", ""],
                    ["", ""],
                    ["EVENT CORE DETAILS", ""],
                    ["EVENT ID", event['id']],
                    ["EVENT TITLE", event['title'] or "Untitled Event"],
                    ["ORGANIZATION", event['club_name'] or "N/A"],
                    ["LEAD ORGANIZER", event['organizer_name'] or "N/A"],
                    ["STATUS", (event['status'] or 'N/A').upper()],
                    ["DESCRIPTION", event['description'] or "No description provided"],
                    ["", ""],
                    ["LOGISTICS & VENUE", ""],
                    ["HALL NAME", event['venue_name'] or "TBA"],
                    ["CAPACITY", event['capacity'] or "N/A"],
                    ["LOCATION", event['venue_description'] or "N/A"],
                    ["ATTENDANCE CODE", event['attendance_code'] or "NOT GENERATED"],
                    ["", ""],
                    ["SCHEDULE & DEADLINES", ""],
                    ["START DATE", event['start_date'].strftime("%Y-%m-%d %H:%M") if event['start_date'] else "N/A"],
                    ["END DATE", event['end_date'].strftime("%Y-%m-%d %H:%M") if event['end_date'] else "N/A"],
                    ["REG DEADLINE", event['reg_deadline'].strftime("%Y-%m-%d %H:%M") if event['reg_deadline'] else "N/A"],
                    ["CREATED AT", event['created_at'].strftime("%Y-%m-%d %H:%M") if event['created_at'] else "N/A"],
                    ["", ""],
                    ["FINANCIALS & TEAMS", ""],
                    ["REGISTRATION FEE", f"INR {event['reg_amount']}" if event['reg_amount'] else "FREE"],
                    ["MIN TEAM SIZE", event['min_team_size'] or 1],
                    ["MAX TEAM SIZE", event['team_size'] or 1],
                    ["FEMALE MANDATORY", "YES" if event['female_mandatory'] else "NO"],
                    ["", ""],
                    ["EVENT FLOW (DETAILED)", ""],
                    [flow_str, ""],
                    ["", ""],
                    ["REFRESHMENTS DATA", ""],
                    [refreshments_str, ""],
                    ["", ""],
                    ["PARTICIPATION METRICS", ""],
                    ["TOTAL REGISTRATIONS", len(participants)],
                    ["MANUAL PRESENT", sum(1 for p in participants if p['manual_attendance'])],
                    ["OTP VERIFIED", sum(1 for p in participants if p['otp_attendance'])]
                ]
                for row in summary_data: ws1.append(row)
                
                # Style Summary
                header_font = Font(bold=True, size=14, color="1E3A8A")
                ws1['A1'].font = header_font
                for i in [3, 11, 17, 23, 29, 32, 35]: # Label headers
                    ws1.cell(row=i, column=1).font = Font(bold=True, size=11)

                # Sheet 2: Participant Register
                ws2 = wb.create_sheet("Participant Register")
                headers = ['STUDENT NAME', 'REG NO', 'EMAIL', 'COLLEGE EMAIL', 'TEAM IDENTITY', 'REG STATUS', 'REG DATE', 'MANUAL PRESENCE', 'OTP VERIFIED']
                ws2.append(headers)
                
                # Style Headers
                h_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
                h_font = Font(bold=True, color="FFFFFF")
                for cell in ws2[1]:
                    cell.fill = h_fill
                    cell.font = h_font
 
                for p in participants:
                    ws2.append([
                        p['full_name'], p['reg_no'], p['email'], p['college_email'],
                        p['team_name'] or "Individual",
                        p['reg_status'], p['registered_at'].replace(tzinfo=None) if p['registered_at'] else "N/A",
                        "PRESENT" if p['manual_attendance'] else "ABSENT",
                        "VERIFIED" if p['otp_attendance'] else "UNVERIFIED"
                    ])

                # Autofit columns
                for sheet in [ws1, ws2]:
                    for col in sheet.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except: pass
                        sheet.column_dimensions[column].width = max_length + 3

                file_stream = io.BytesIO()
                wb.save(file_stream)
                file_stream.seek(0)
                
                # Sanitize filename for headers
                clean_title = "".join(c for c in (event['title'] or str(event_id)) if c.isalnum() or c in (' ', '_')).replace(' ', '_')
                filename = f"Dossier_{clean_title}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                return send_file(
                    file_stream,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name=filename
                )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
@admin_bp.route('/sessions/all', methods=['GET'])
@require_auth(roles=['admin'])
def get_all_sessions(current_user):
    """Admin only: Monitor all active and past session activities."""
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Fetch recent activities for all users
                cur.execute("""
                    SELECT s.action, s.ip_address, s.user_agent, s.timestamp, u.full_name, u.email, u.role
                    FROM user_session_history s
                    JOIN users u ON s.user_id = u.id
                    ORDER BY s.timestamp DESC
                    LIMIT 100
                """)
                return jsonify(cur.fetchall()), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
