from flask import Blueprint, request, jsonify
from utils.auth_utils import require_auth
from db import DatabaseConnection
from psycopg2.extras import RealDictCursor
import razorpay
from utils.crypto_utils import decrypt_data
from utils.invoice_generator import generate_and_send_invoice, generate_and_send_cultural_ticket
from datetime import datetime, timedelta

culturals_bp = Blueprint('culturals', __name__)

@culturals_bp.route('/list', methods=['GET'])
@require_auth(roles=['student', 'organizer', 'admin'])
def get_culturals(current_user):
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT c.*, cl.name as club_name,
                           (SELECT COUNT(*) FROM cultural_bookings WHERE cultural_id = c.id AND status = 'confirmed') as tickets_sold,
                           EXISTS(SELECT 1 FROM cultural_bookings WHERE cultural_id = c.id AND student_id = %s AND status = 'confirmed') as user_booked
                    FROM culturals c
                    JOIN clubs cl ON c.club_id = cl.id
                    ORDER BY c.event_date ASC
                """, (int(current_user['sub']),))
                return jsonify(cur.fetchall()), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@culturals_bp.route('/create', methods=['POST'])
@require_auth(roles=['organizer', 'admin'])
def create_cultural(current_user):
    data = request.json
    
    event_date_str = data.get('event_date')
    booking_deadline_str = data.get('booking_deadline')
    
    if event_date_str and booking_deadline_str:
        try:
            # Clean string for naive parsing (remove ms, Z, or timezone offsets)
            e_str = event_date_str.split('.')[0].replace('Z', '').split('+')[0]
            b_str = booking_deadline_str.split('.')[0].replace('Z', '').split('+')[0]
            
            event_date = datetime.fromisoformat(e_str)
            booking_deadline = datetime.fromisoformat(b_str)
            now = datetime.now()
            
            if event_date < now:
                return jsonify({"error": "Event date cannot be in the past"}), 400
            if booking_deadline < now:
                return jsonify({"error": "Booking deadline cannot be in the past"}), 400
            if booking_deadline >= event_date:
                return jsonify({"error": "Booking deadline must be strictly before the event date"}), 400
        except ValueError:
            return jsonify({"error": "Invalid date format"}), 400

    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Get organizer's club
                cur.execute("SELECT club_id FROM users WHERE id = %s", (int(current_user['sub']),))
                user = cur.fetchone()
                if not user or not user['club_id']:
                    return jsonify({"error": "No club associated with this account"}), 403
                
                cur.execute("""
                    INSERT INTO culturals (title, description, price, total_tickets, available_tickets, event_date, booking_deadline, venue, club_id, template_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (data['title'], data.get('description'), data['price'], data['total_tickets'], data['total_tickets'], data.get('event_date'), data.get('booking_deadline'), data.get('venue'), user['club_id'], data.get('template_id', 'classic_purple')))
                conn.commit()
                return jsonify({"message": "Cultural unit created successfully"}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@culturals_bp.route('/update-tickets', methods=['POST'])
@require_auth(roles=['organizer', 'admin'])
def update_tickets(current_user):
    data = request.json
    cultural_id = data.get('cultural_id')
    new_total = int(data.get('total_tickets'))
    
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT total_tickets, available_tickets FROM culturals WHERE id = %s", (cultural_id,))
                cult = cur.fetchone()
                if not cult: return jsonify({"error": "Cultural unit not found"}), 404
                
                if new_total < cult['total_tickets']:
                    return jsonify({"error": "Ticket quantity can only be increased, not decreased."}), 400
                
                diff = new_total - cult['total_tickets']
                cur.execute("""
                    UPDATE culturals 
                    SET total_tickets = %s, available_tickets = available_tickets + %s 
                    WHERE id = %s
                """, (new_total, diff, cultural_id))
                conn.commit()
                return jsonify({"message": f"Tickets increased by {diff}. Total is now {new_total}."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@culturals_bp.route('/book', methods=['POST'])
@require_auth(roles=['student'])
def book_ticket(current_user):
    data = request.json
    cultural_id = data.get('cultural_id')
    student_id = int(current_user['sub'])
    
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # 1. Check if already booked
                cur.execute("SELECT id FROM cultural_bookings WHERE cultural_id = %s AND student_id = %s", (cultural_id, student_id))
                if cur.fetchone():
                    return jsonify({"error": "You have already booked a ticket for this cultural event."}), 400
                
                # 2. Check availability
                cur.execute("SELECT c.*, cl.name as club_name, cl.razorpay_key_id, cl.razorpay_key_secret FROM culturals c JOIN clubs cl ON c.club_id = cl.id WHERE c.id = %s", (cultural_id,))
                cult = cur.fetchone()
                if not cult: return jsonify({"error": "Cultural unit not found"}), 404
                
                # 3. Check Deadline
                now = datetime.now()
                if cult['event_date'] and now > cult['event_date']:
                    return jsonify({"error": "Cannot book a ticket for a past event."}), 400
                if cult['booking_deadline'] and now > cult['booking_deadline']:
                    return jsonify({"error": "Registration for this event has closed."}), 400

                if cult['available_tickets'] <= 0:
                    return jsonify({"error": "Event Full!"}), 400
                
                price = float(cult['price'])
                
                if price <= 0:
                    # FREE TICKET
                    cur.execute("""
                        INSERT INTO cultural_bookings (cultural_id, student_id, status, amount_paid, ticket_id)
                        VALUES (%s, %s, 'confirmed', 0, %s)
                        RETURNING id
                    """, (cultural_id, student_id, "TBD")) # Will update below
                    recorded_booking_id = cur.fetchone()['id']
                    ticket_str = f"CULT-{recorded_booking_id}"
                    cur.execute("UPDATE cultural_bookings SET ticket_id = %s WHERE id = %s", (ticket_str, recorded_booking_id))
                    cur.execute("UPDATE culturals SET available_tickets = available_tickets - 1 WHERE id = %s", (cultural_id,))
                    
                    # Send Invoice/Ticket
                    cur.execute("SELECT full_name, email, college_email, reg_no FROM users WHERE id = %s", (student_id,))
                    u = cur.fetchone()
                    # Generate Both: Ticket and Invoice
                    ticket_path = generate_and_send_cultural_ticket(
                        u['full_name'], [u['email'], u['college_email']], cult['title'], cult['club_name'], 0, f"CULT-{recorded_booking_id}", 
                        (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime("%Y-%m-%d %H:%M"), u['reg_no'], cult['venue'], cult.get('template_id', 'classic_purple'),
                        send_email=False 
                    )
                    invoice_path = generate_and_send_invoice(
                        u['full_name'], [u['email'], u['college_email']], cult['title'], cult['club_name'], 0, "FREE_CULT", 
                        (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime("%Y-%m-%d %H:%M"), u['reg_no'], send_email=False
                    )

                    # Send Unified Email
                    from utils.invoice_generator import send_combined_email
                    send_combined_email(
                        u['full_name'], [u['email'], u['college_email']], cult['title'], [ticket_path, invoice_path]
                    )
                    
                    conn.commit()
                    return jsonify({"message": "Ticket booked successfully! Check your email."}), 200
                else:
                    # PAID TICKET: Razorpay
                    dec_key_id = decrypt_data(cult['razorpay_key_id']) if cult.get('razorpay_key_id') else None
                    dec_key_secret = decrypt_data(cult['razorpay_key_secret']) if cult.get('razorpay_key_secret') else None
                    if not dec_key_id or not dec_key_secret:
                        return jsonify({"error": "Payment gateway not configured for this club"}), 400
                    
                    client = razorpay.Client(auth=(dec_key_id, dec_key_secret))
                    order = client.order.create(dict(amount=int(price * 100), currency='INR', receipt=f"cult_{cultural_id}_{student_id}"))
                    
                    cur.execute("""
                        INSERT INTO cultural_bookings (cultural_id, student_id, status, razorpay_order_id, amount_paid, payment_initiated_at)
                        VALUES (%s, %s, 'pending', %s, %s, CURRENT_TIMESTAMP)
                    """, (cultural_id, student_id, order['id'], price))
                    conn.commit()
                    return jsonify({
                        "type": "paid", "order_id": order['id'], "amount": price, "key": dec_key_id
                    }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@culturals_bp.route('/verify-booking', methods=['POST'])
@require_auth(roles=['student'])
def verify_booking(current_user):
    data = request.json
    razorpay_payment_id = data.get('razorpay_payment_id')
    razorpay_order_id = data.get('razorpay_order_id')
    razorpay_signature = data.get('razorpay_signature')
    
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT cb.*, c.title, c.venue, c.club_id, c.template_id, cl.name as club_name, cl.razorpay_key_id, cl.razorpay_key_secret,
                           u.full_name, u.email, u.college_email, u.reg_no
                    FROM cultural_bookings cb
                    JOIN culturals c ON cb.cultural_id = c.id
                    JOIN clubs cl ON c.club_id = cl.id
                    JOIN users u ON cb.student_id = u.id
                    WHERE cb.razorpay_order_id = %s AND cb.status = 'pending'
                """, (razorpay_order_id,))
                booking = cur.fetchone()
                if not booking: return jsonify({"error": "Booking record not found"}), 404
                
                dec_key_id = decrypt_data(booking['razorpay_key_id'])
                dec_key_secret = decrypt_data(booking['razorpay_key_secret'])
                client = razorpay.Client(auth=(dec_key_id, dec_key_secret))
                
                # Verify signature
                client.utility.verify_payment_signature({
                    'razorpay_order_id': razorpay_order_id,
                    'razorpay_payment_id': razorpay_payment_id,
                    'razorpay_signature': razorpay_signature
                })
                
                # Update booking
                cur.execute("""
                    UPDATE cultural_bookings 
                    SET status = 'confirmed', razorpay_payment_id = %s 
                    WHERE id = %s
                """, (razorpay_payment_id, booking['id']))
                
                # Deduct ticket
                cur.execute("UPDATE culturals SET available_tickets = available_tickets - 1 WHERE id = %s", (booking['cultural_id'],))
                
                # Generate Both Documents (Using Database ID as Ticket ID)
                ticket_str = f"CULT-{booking['id']}"
                cur.execute("UPDATE cultural_bookings SET ticket_id = %s WHERE id = %s", (ticket_str, booking['id']))
                
                ticket_path = generate_and_send_cultural_ticket(
                    booking['full_name'], [booking['email'], booking['college_email']], booking['title'], booking['club_name'], 
                    booking['amount_paid'], ticket_str, (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime("%Y-%m-%d %H:%M"), booking['reg_no'], booking['venue'], booking.get('template_id', 'classic_purple'),
                    send_email=False
                )
                invoice_path = generate_and_send_invoice(
                    booking['full_name'], [booking['email'], booking['college_email']], booking['title'], booking['club_name'], 
                    booking['amount_paid'], razorpay_payment_id, (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime("%Y-%m-%d %H:%M"), booking['reg_no'], send_email=False
                )

                from utils.invoice_generator import send_combined_email
                send_combined_email(
                    booking['full_name'], [booking['email'], booking['college_email']], booking['title'], [ticket_path, invoice_path]
                )
                
                conn.commit()
                return jsonify({"message": "Ticket booked successfully! Check your email."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@culturals_bp.route('/bookings', methods=['GET'])
@require_auth(roles=['organizer', 'admin'])
def get_cultural_bookings(current_user):
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT club_id FROM users WHERE id = %s", (int(current_user['sub']),))
                user = cur.fetchone()
                
                cur.execute("""
                    SELECT cb.id, cb.status, cb.booked_at, cb.amount_paid,
                           c.title as cultural_title,
                           u.full_name as student_name, u.reg_no, u.email, u.college_email
                    FROM cultural_bookings cb
                    JOIN culturals c ON cb.cultural_id = c.id
                    JOIN users u ON cb.student_id = u.id
                    WHERE c.club_id = %s
                    ORDER BY cb.booked_at DESC
                """, (user['club_id'],))
                return jsonify(cur.fetchall()), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@culturals_bp.route('/delete/<int:cultural_id>', methods=['DELETE'])
@require_auth(roles=['admin', 'organizer'])
def delete_cultural(current_user, cultural_id):
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Permission check
                cur.execute("SELECT club_id FROM culturals WHERE id = %s", (cultural_id,))
                cult = cur.fetchone()
                if not cult: return jsonify({"error": "Cultural unit not found"}), 404
                
                # Check if current user is admin of the same organization or the organizer of the same club
                # For admin, we should check organization_name/club_id match if we want to be strict, 
                # but standard practice in this app seems to be global admin or club-based organizer.
                if current_user['role'] == 'organizer':
                     cur.execute("SELECT club_id FROM users WHERE id = %s", (int(current_user['sub']),))
                     user_club = cur.fetchone()
                     if not user_club or user_club['club_id'] != cult['club_id']:
                         return jsonify({"error": "Unauthorized"}), 403
                
                cur.execute("DELETE FROM culturals WHERE id = %s", (cultural_id,))
                conn.commit()
                return jsonify({"message": "Cultural unit deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@culturals_bp.route('/my-bookings', methods=['GET'])
@require_auth(roles=['student'])
def get_my_cultural_bookings(current_user):
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Cleanup: Automatically cancel pending cultural bookings that are > 5 minutes old
                cur.execute("""
                    UPDATE cultural_bookings 
                    SET status = 'cancelled' 
                    WHERE status = 'pending' 
                      AND payment_initiated_at < CURRENT_TIMESTAMP - INTERVAL '5 minutes'
                """)
                conn.commit()
                
                cur.execute("""
                    SELECT cb.id, cb.status, cb.booked_at, cb.amount_paid, cb.razorpay_payment_id, cb.payment_initiated_at,
                           c.title as cultural_title, c.venue, c.event_date
                    FROM cultural_bookings cb
                    JOIN culturals c ON cb.cultural_id = c.id
                    WHERE cb.student_id = %s
                    ORDER BY cb.booked_at DESC
                """, (int(current_user['sub']),))
                return jsonify(cur.fetchall()), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
from flask import Blueprint, request, jsonify, Response, send_file
import io
import openpyxl
from openpyxl.styles import Font, PatternFill

@culturals_bp.route('/export/<int:cultural_id>', methods=['GET'])
@require_auth(roles=['admin', 'organizer'])
def export_cultural_bookings(current_user, cultural_id):
    """Export the list of attendees for a cultural unit as a true Excel (.xlsx) file."""
    try:
        with DatabaseConnection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT id, title, club_id FROM culturals WHERE id = %s", (cultural_id,))
                cult = cur.fetchone()
                if not cult: return jsonify({"error": "Event not found"}), 404
                
                if current_user['role'] == 'organizer' and str(cult['club_id']) != str(current_user.get('club_id')):
                    return jsonify({"error": "Unauthorized"}), 403

                cur.execute("""
                    SELECT u.full_name, u.reg_no, u.email, u.college_email, 
                           cb.ticket_id, cb.status, cb.booked_at, cb.amount_paid
                    FROM cultural_bookings cb
                    JOIN users u ON cb.student_id = u.id
                    WHERE cb.cultural_id = %s
                    ORDER BY cb.booked_at ASC
                """, (cultural_id,))
                bookings = cur.fetchall()

                # Generate XLSX using openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Attendees"

                # Define headers and styling
                headers = ['STUDENT NAME', 'REG NO', 'EMAIL', 'COLLEGE EMAIL', 'TICKET ID', 'STATUS', 'BOOKED AT', 'AMOUNT PAID']
                ws.append(headers)

                # Style header
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # CampusHub Dark Blue
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill

                # Add Data
                for b in bookings:
                    ws.append([
                        b['full_name'], b['reg_no'], b['email'], b['college_email'],
                        b['ticket_id'], b['status'], b['booked_at'].replace(tzinfo=None),
                        f"INR {b['amount_paid']}"
                    ])

                # Adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column].width = max_length + 2

                file_stream = io.BytesIO()
                wb.save(file_stream)
                file_stream.seek(0)
                
                # Sanitize filename for headers
                clean_title = "".join(c for c in (cult['title'] or str(cultural_id)) if c.isalnum() or c in (' ', '_')).replace(' ', '_')
                filename = f"Attendees_{clean_title}_{datetime.now().strftime('%Y%m%d')}.xlsx"

                return send_file(
                    file_stream,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name=filename
                )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
